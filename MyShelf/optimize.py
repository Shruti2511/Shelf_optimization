import tkinter as tk
from tkinter import ttk
import matplotlib.pyplot as plt
import openpyxl
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg)
import numpy as np
import pandas as pd
import statsmodels.api as sm
from statsmodels.tsa.arima_model import ARIMA
import warnings
warnings.filterwarnings("ignore")
from pmdarima import auto_arima
import datetime
import math

months = ["jan", "feb", "march", "apr", "may", "june", "july", "aug", "sept", "oct", "nov", "dec"]

#frame_styles = {"relief": "groove", "bd": 3, "bg": "#BEB2A7", "fg": "#073bb3", "font": ("Arial", 9, "bold")}

class LoginPage(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        main_frame = tk.Frame(self, height=431, width=626)  # this is the background
        main_frame.pack(fill="both", expand="true")

        self.geometry("400x180")  # Sets window size to 626w x 431h pixels
        self.resizable(0, 0)  # This prevents any resizing of the screen

        frame_login = tk.Frame(main_frame, relief="groove")  # this is the frame that holds all the login details and buttons
        frame_login.place(rely=0, relx=0.05, height=160, width=400)

        label_title = tk.Label(frame_login, text="Login Page", anchor="center")
        label_title.grid(row=0, column=1, columnspan=1, sticky="news", padx=10, pady=10)

        label_user = tk.Label(frame_login, text="Username:", anchor="center")
        label_user.grid(row=1, column=0, sticky="news", padx=5, pady=5)

        label_pw = tk.Label(frame_login, text="Password:", anchor="center")
        label_pw.grid(row=2, column=0, sticky="news", padx=5, pady=5)

        entry_user = ttk.Entry(frame_login, width=45, cursor="xterm")
        entry_user.grid(row=1, column=1, sticky="news", padx=5, pady=5)

        entry_pw = ttk.Entry(frame_login, width=45, cursor="xterm", show="*")
        entry_pw.grid(row=2, column=1, sticky="news", padx=5, pady=5)

        blank1 = tk.Label(frame_login, text="", anchor="center")
        blank1.grid(row=3, columnspan=2, sticky="news", padx=5, pady=5)

        blank = tk.Label(frame_login, text="", anchor="center")
        blank.grid(row=4, columnspan=2, sticky="news", padx=5, pady=5)

        button = ttk.Button(frame_login, text="Login", command=lambda: getlogin())
        button.place(rely=0.68, relx=0.30)

        signup_btn = ttk.Button(frame_login, text="Register", command=lambda: get_signup())
        signup_btn.place(rely=0.68, relx=0.50)

        def get_signup():
            register()

        def getlogin():
            username = entry_user.get()
            password = entry_pw.get()
            # if your want to run the script as it is set validation = True
            validation = validate(username, password)
            if validation:
                blank.config(text="")
                #tk.messagebox.showinfo("Login Successful","Welcome {}".format(username))
                root.deiconify()
                top.destroy()
            else:
                blank.config(text="The Username or Password you have entered are incorrect")

        def validate(username, password):
            # Checks the text file for a username/password combination.
            try:
                with open("credentials.txt", "r") as credentials:
                    for line in credentials:
                        line = line.split(",")
                        if line[1] == username and line[3] == password:
                            return True
                    return False
            except FileNotFoundError:
                blank.config(text="You need to Register first")
                return False

class register(tk.Tk):

    def __init__(self, *args, **kwargs):

        tk.Tk.__init__(self, *args, **kwargs)

        register_frame = tk.Frame(self, height=150, width=250)
        # pack_propagate prevents the window resizing to match the widgets
        register_frame.pack_propagate(0)
        register_frame.pack(fill="both", expand="true")
        self.geometry("400x180")
        self.resizable(0, 0)

        self.title("Registration")
        
        label_title = tk.Label(register_frame, text="Registration", anchor="center")
        label_title.grid(row=0, columnspan=2, sticky="news", padx=10, pady=10)

        label_user = tk.Label(register_frame, width=20, text="New Username:")
        label_user.grid(row=1, column=0, sticky="news", padx=5, pady=5)

        blank1 = tk.Label(register_frame, text="", anchor="center")
        blank1.grid(row=3, columnspan=2, sticky="news", padx=5, pady=5)

        blank = tk.Label(register_frame, text="", anchor="center")
        blank.grid(row=4, columnspan=2, sticky="news", padx=5, pady=5)

        label_pw = tk.Label(register_frame, width=20, text="New Password:")
        label_pw.grid(row=2, column=0, sticky="news", padx=5, pady=5)

        entry_user = ttk.Entry(register_frame, width=30, cursor="xterm")
        entry_user.grid(row=1, column=1, sticky="news", padx=5, pady=5)

        entry_pw = ttk.Entry(register_frame, width=30, cursor="xterm", show="*")
        entry_pw.grid(row=2, column=1, sticky="news", padx=5, pady=5)

        button = ttk.Button(register_frame, width=15, text="Reset Password", command=lambda: reset())
        button.place(rely=0.60, relx=0.22)

        button = ttk.Button(register_frame, width=15, text="Create Account", command=lambda: signup())
        button.place(rely=0.60, relx=0.50)

        def reset():
            user = entry_user.get()
            pw = entry_pw.get()
            i=-1
            if len(pw)>3:
                
                try:
                    with open("credentials.txt", "r") as credentials:
                        data = credentials.readlines()
                        for n in data:
                            n = n.split(",")
                            i=i+1
                            if n[1] == user:
                                data[i]=(data[i][:-len(n[3])-2]+pw+",\n")
                                with open("credentials.txt", "w") as credentials:
                                    credentials.writelines(data)
                                blank.config(text="Password Updated")
                                break
                            else:
                                blank.config(text="User Doesn't Exists")

                except FileNotFoundError:
                    blank.config(text="You need to Register first")

            else:
                blank.config(text="Password should be longer than 3 values")            

        def signup():
            # Creates a text file with the Username and password
            user = entry_user.get()
            pw = entry_pw.get()
            validation = validate_user(user)
            if user=="":
                blank.config(text="Username cannot be blank")
            else:
                if not validation:
                    blank.config(text="Username already exists")
                else:
                    if len(pw) > 3:
                        credentials = open("credentials.txt", "a")
                        credentials.write(f"Username,{user},Password,{pw},\n")
                        credentials.close()
                        blank.config(text="Your account details have been stored")
                        register.destroy(self)

                    else:
                        blank.config(text="Password should be longer than 3 values")

        def validate_user(username):
            # Checks the text file for a username/password combination.
            try:
                with open("credentials.txt", "r") as credentials:
                    for line in credentials:
                        line = line.split(",")
                        if line[1] == username:
                            return False
                return True
            except FileNotFoundError:
                return True

class MenuBar(tk.Menu):
    def __init__(self, parent):
        tk.Menu.__init__(self, parent)

        self.add_cascade(label="Products", command=lambda: parent.show_frame(products))

        self.add_cascade(label="Pricing", command=lambda: parent.show_frame(pricing))

        self.add_cascade(label="Forecast",  command=lambda: parent.show_frame(Forecast))

        self.add_cascade(label="Optimize", command=lambda: parent.show_frame(optimize))

        menu_help = tk.Menu(self, tearoff=0)
        self.add_cascade(label="Help", menu=menu_help)
        menu_help.add_command(label="About", command=lambda: parent.show_frame(About))
        menu_help.add_separator()
        menu_help.add_command(label="Exit Application", command=lambda: parent.Quit_application())

class MyApp(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        main_frame = tk.Frame(self, height=600, width=1024)
        main_frame.pack_propagate(0)
        main_frame.pack(fill="both", expand="true")
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)
        self.resizable(0, 0) #prevents the app from being resized

        self.frames = {}
        pages = (products, optimize, pricing, Forecast, About)
        for F in pages:
            frame = F(main_frame, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame(products)
        menubar = MenuBar(self)
        tk.Tk.config(self, menu=menubar)

    def show_frame(self, name):
        frame = self.frames[name]
        frame.tkraise()

    def OpenNewWindow(self):
        OpenNewWindow()

    def Quit_application(self):
        self.destroy()

class GUI(tk.Frame):
    def __init__(self, parent):
        tk.Frame.__init__(self, parent)
        self.main_frame = tk.Frame(self, height=500, width=800)
        self.main_frame.pack(fill="both", expand="true")
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)

class products(GUI):  # inherits from the GUI class
    def __init__(self, parent, controller):
        GUI.__init__(self, parent)

        df = pd.read_excel('data.xlsx')

        frame1 = tk.LabelFrame(self, text="Product Details")
        frame1.place(rely=0.06, relx=0.02, height=455, width=500)

        tv1 = ttk.Treeview(frame1)
        column_list = ["Sr no", "Name", "Profit", "Volume", "Last Month sales"]
        tv1['columns'] = column_list
        tv1["show"] = "headings"  # removes empty column
        for column in column_list:
            tv1.heading(column, text=column)
            tv1.column(column, anchor="center")
        tv1.column(0, width=50, minwidth=50)
        tv1.column(1, width=150, minwidth=150, anchor="w")
        tv1.column(2, width=80, minwidth=80)
        tv1.column(3, width=80, minwidth=80)
        tv1.column(4, width=120, minwidth=120)
        tv1.place(relheight=1, relwidth=0.995)
        treescroll = tk.Scrollbar(frame1)
        treescroll.configure(command=tv1.yview)
        tv1.configure(yscrollcommand=treescroll.set)
        treescroll.pack(side="right", fill="y")

        blank = tk.Label(self)
        blank.place(rely=0.02, relx=0.02)
        blank.config(text="WELCOME")

        product_category = ["Select All"]
        value_inside = tk.StringVar(self)
        product_category_select = []
         
        def option_selected(selection):
            tv1.delete(*tv1.get_children())
            df = pd.read_excel('data.xlsx')
            product_volume = df['product_dimL']*df['product_dimB']*df['product_dimH']
            if selection == "Select All":
                for i in range(len(df["product_name"])):
                    tv1.insert('','end', values=(i+1, df["product_name"][i], float(df['product_SP'][i]) - float(df['product_CP'][i]), product_volume[i]))
            
            else:
                df = pd.read_excel('data.xlsx')
                product_volume = df['product_dimL']*df['product_dimB']*df['product_dimH']
                for i in range(len(df['product_name'])):    
                    if selection == df['category'][i]:
                        tv1.insert('','end', values=(i+1, df["product_name"][i], float(df['product_SP'][i]) - float(df['product_CP'][i]), product_volume[i]))

        c = pd.read_excel('data.xlsx', sheet_name="category")
        for c in c['Product Category']:
            product_category.append(c)
            product_category_select.append(c)
        
        category = ttk.OptionMenu(self, value_inside, "Select Category", *product_category, command=option_selected)
        category.place(rely=0.02, relx=0.23)
        category.config(width=24)

        def addRecord():
            df = pd.read_excel('data.xlsx')
            if product_name_entry.get() != "" and dimL_entry.get()!="" and dimB_entry.get()!="" and dimH_entry.get()!="" and product_CPprice_entry.get()!="" and product_SPprice_entry.get()!="":
                try:
                    float(dimL_entry.get()) and float(dimB_entry.get()) and float(dimH_entry.get()) and float(product_CPprice_entry.get()) and float(product_SPprice_entry.get())

                    product_name =df['product_name'].tolist()
                    if product_name_entry.get() not in product_name:
                        workbook = openpyxl.load_workbook('data.xlsx')
                        worksheet = workbook.active
                        worksheet.append([float(product_name_entry.get()),float(prod_selection.get()),float(dimL_entry.get()),float(dimB_entry.get()),float(dimH_entry.get()),float(product_CPprice_entry.get()),float(product_SPprice_entry.get())])
                        workbook.save('data.xlsx')    
                        workbook.close()
                        tv1.delete(*tv1.get_children())
                        df = pd.read_excel('data.xlsx')
                        product_volume = df['product_dimL']*df['product_dimB']*df['product_dimH']
                        for n in range(len(df['product_name'])):
                            tv1.insert('','end', values=(n+1, df['product_name'][n], df['product_CPprofit'][n], df['product_SPprofit'][n], product_volume[n]))
                        product_name_entry.delete(0, tk.END)
                        product_CPprice_entry.delete(0, tk.END)
                        product_SPprice_entry.delete(0, tk.END)
                        dimL_entry.delete(0, tk.END)
                        dimB_entry.delete(0, tk.END)
                        dimH_entry.delete(0, tk.END)
                        blank.config(text="Product Added Successfully")
                    else:
                        blank.config(text="Product Name Already exists")

                except(ValueError):
                    blank.config(text="Please Enter Required Format")
            else:
                blank.config(text="Please Enter Required Values")
            value_inside.set("Select Category")

        addRecord_button = tk.Button(self, text="Add Record", command = addRecord)
        addRecord_button.place(rely=0.02, relx=0.89, height=20)

        def search():
            search_text = search_box.get()
            if search_text != "":
                for item in tv1.get_children():
                    if search_text.lower() in str(tv1.item(item)["values"][1]).lower():
                        tv1.item(item, open=True)
                        tv1.selection_set(item)
                    else:
                        tv1.item(item, open=False)
                        tv1.selection_remove(item)
            else:
                blank.config(text="Please Enter Value to Search")

        search_box = tk.Entry(self)
        search_box.place(rely=0.02, relx=0.45, height=20)
        search_box.bind("<KeyRelease>", search)
        
        search_button = tk.Button(self, command = search)
        search_button.place(rely=0.02, relx=0.61, height=20, width=22)

        tv1.delete(*tv1.get_children())
        product_name = df['product_name']
        product_volume = df['product_dimL']*df['product_dimB']*df['product_dimH']
        for n in range(len(product_name)):
            tv1.insert('','end', values=(n+1, product_name[n], df['product_SP'][n]-df['product_CP'][n], product_volume[n], df[df.keys()[-1]][n]))

        frame2 = tk.LabelFrame(self, text="Sales Details")
        frame2.place(rely=0.06, relx=0.65, height=455, width=280)

        product_name_label = tk.Label(frame2, text="Name", width=12)
        product_name_label.grid(row=0, column=0, padx=1, pady=5, sticky="news")
        product_name_entry = tk.Entry(frame2, width=25)
        product_name_entry.grid(row=0, column=1, padx=10, pady=5, sticky="news")
        product_category_label = tk.Label(frame2, text="Category", width=12)
        product_category_label.grid(row=1, column=0, padx=1, pady=5, sticky="news")
        prod_selection = tk.StringVar(frame2)
        product_category_menu = ttk.OptionMenu(frame2, prod_selection, "Select Category", *product_category_select)
        product_category_menu.grid(row=1, column=1, padx=10, pady=5, sticky="news")
        product_CPprice_label = tk.Label(frame2, text="Cost Price")
        product_CPprice_label.grid(row=2, column=0, padx=1, pady=5, sticky="news")
        product_CPprice_entry = tk.Entry(frame2)
        product_CPprice_entry.grid(row=2, column=1, padx=10, pady=5, sticky="news")
        product_SPprice_label = tk.Label(frame2, text="Selling Price")
        product_SPprice_label.grid(row=3, column=0, padx=1, pady=5, sticky="news")
        product_SPprice_entry = tk.Entry(frame2)
        product_SPprice_entry.grid(row=3, column=1, padx=10, pady=5, sticky="news")
        
        dimL_label = tk.Label(frame2, text="Length")
        dimL_label.grid(row=4, column=0, padx=5, pady=1, sticky="news")
        dimL_entry = tk.Entry(frame2)
        dimL_entry.grid(row=4, column=1, padx=10, pady=5, sticky="news")
        dimB_label = tk.Label(frame2, text="Breadth")
        dimB_label.grid(row=5, column=0, padx=1, pady=5, sticky="news")
        dimB_entry = tk.Entry(frame2)
        dimB_entry.grid(row=5, column=1, padx=10, pady=5, sticky="news")
        dimH_label = tk.Label(frame2, text="Height")
        dimH_label.grid(row=6, column=0, padx=1, pady=5, sticky="news")
        dimH_entry = tk.Entry(frame2)
        dimH_entry.grid(row=6, column=1, padx=10, pady=5, sticky="news")

        blank1 = tk.Label(frame2) 
        blank1.grid(row=7, columnspan=2, padx=1, pady=5, sticky="news")
        sales = tk.Label(frame2, text="Sales", relief="solid")
        sales.grid(row=8, columnspan=2, padx=1, pady=5, sticky="news")
        month_label = tk.Label(frame2, text="Month", width=10)
        month_label.grid(row=9, column=0, padx=1, pady=5, sticky="news")
        month_option_menu = ttk.OptionMenu(frame2, tk.StringVar(), *["","January", "Feb", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
        month_option_menu.grid(row=9, column=1, padx=1, pady=5, sticky="news")
        year_label = tk.Label(frame2, text="Year")
        year_label.grid(row=10, column=0, padx=1, pady=5, sticky="news")
        current_year = datetime.datetime.now().year
        years = list(range(2021, current_year + 1))
        years_menu = ttk.OptionMenu(frame2, tk.StringVar(), *years)
        years_menu.grid(row=10, column=1, padx=1, pady=5, sticky="news")
        sales_label = tk.Label(frame2, text="Sales")
        sales_label.grid(row=11, column=0, padx=1, pady=5, sticky="news")
        sales_value = tk.Entry(frame2)
        sales_value.grid(row=11, column=1, padx=10, pady=5, sticky="news")
        add_sales_button = tk.Button(frame2, text="Add Sales")
        add_sales_button.grid(row=12, columnspan=2, padx=1, pady=5, sticky="news")

        def addNew():
            product_name_entry.delete(0, tk.END)
            prod_selection.set("Select Category")
            product_CPprice_entry.delete(0, tk.END)
            product_SPprice_entry.delete(0, tk.END)
            dimL_entry.delete(0, tk.END)
            dimB_entry.delete(0, tk.END)
            dimH_entry.delete(0, tk.END)        
            frame2.config(text="Sales Details")
            value_inside.set("Select Category")
            blank.config(text="")
            tv1.selection_remove(tv1.selection())
            addNew_button.place_forget()

        addNew_button = tk.Button(self, text=" Add New  ", command = addNew, width=9)
        addNew_button.place(rely=0.02, relx=0.89, height=20)

        def on_select(event):
            df = pd.read_excel('data.xlsx')
            selected_item = tv1.item(tv1.selection()[0], 'values')[1]
            selected_item_index = int(tv1.item(tv1.selection()[0], 'values')[0]) - 1
            addNew_button.place(rely=0.02, relx=0.89, height=20)
            prod_selection.set(df["category"][selected_item_index])
            product_name_entry.delete(0, tk.END)
            product_name_entry.insert(0, str(selected_item))
            product_CPprice_entry.delete(0, tk.END)
            product_CPprice_entry.insert(0, df['product_CP'][selected_item_index])
            product_SPprice_entry.delete(0, tk.END)
            product_SPprice_entry.insert(0, df['product_SP'][selected_item_index])
            dimL_entry.delete(0, tk.END)
            dimL_entry.insert(0, df["product_dimL"][selected_item_index])
            dimB_entry.delete(0, tk.END)
            dimB_entry.insert(0, df["product_dimB"][selected_item_index])
            dimH_entry.delete(0, tk.END)
            dimH_entry.insert(0, df["product_dimH"][selected_item_index])
            frame2.config(text="Sales Details:" + " " +selected_item)

            def edit():
                workbook = openpyxl.load_workbook('data.xlsx')
                worksheet = workbook.active
                productName = worksheet.cell(row=selected_item_index+2, column=1)
                productName.value = product_name_entry.get()
                productCategory = worksheet.cell(row=selected_item_index+2, column=2)
                productCategory.value = product_category_menu.get()
                dimL = worksheet.cell(row=selected_item_index+2, column=3)
                dimL.value = dimL_entry.get()
                dimB = worksheet.cell(row=selected_item_index+2, column=4)
                dimB.value = dimB_entry.get()
                dimH = worksheet.cell(row=selected_item_index+2, column=5)
                dimH.value = dimH_entry.get()
                prod_CP = worksheet.cell(row=selected_item_index+2, column=6)
                prod_CP.value = product_CPprice_entry.get()
                prod_SP = worksheet.cell(row=selected_item_index+2, column=7)
                prod_SP.value = product_SPprice_entry.get() 
                
                
                workbook.save('data.xlsx')
            
                tv1.delete(*tv1.get_children())
                df = pd.read_excel('data.xlsx')
                product_volume = df['product_dimL']*df['product_dimB']*df['product_dimH']
                for n in range(len(product_volume)):
                    tv1.insert('','end', values=(n+1, df['product_name'][n], df['product_SP'][n]-df['product_CP'][n], product_volume[n]))

                product_name_entry.delete(0, tk.END)
                product_CPprice_entry.delete(0, tk.END)
                product_SPprice_entry.delete(0, tk.END)
                dimL_entry.delete(0, tk.END)
                dimB_entry.delete(0, tk.END)
                dimH_entry.delete(0, tk.END)
                addNew_button.destroy()
                value_inside.set("Select Category")
                frame2.config(text="Sales Details")
                blank.config(text="Record Edited Successfully")
                
            def delete():            
                workbook = openpyxl.load_workbook('data.xlsx')
                worksheet = workbook.active
                worksheet.delete_rows(selected_item_index+2, 1)
                workbook.save('data.xlsx')

                tv1.delete(*tv1.get_children())
                df = pd.read_excel('data.xlsx')
                product_volume = df['product_dimL']*df['product_dimB']*df['product_dimH']
                for n in range(len(product_volume)):
                    tv1.insert('','end', values=(n+1, df['product_name'][n], float(df['product_SP'][n])-float(df['product_CP'][n]), product_volume[n]))
                product_name_entry.delete(0, tk.END)
                product_CPprice_entry.delete(0, tk.END)
                product_SPprice_entry.delete(0, tk.END)
                dimL_entry.delete(0, tk.END)
                dimB_entry.delete(0, tk.END)
                dimH_entry.delete(0, tk.END)
                addNew_button.destroy()
                value_inside.set("Select Category")
                frame2.config(text="Sales Details")
                blank.config(text="Record Deleted Successfully")                
        
            edit = tk.Button(self, text="Edit", command = edit, width=10)
            edit.place(rely=0.02, relx=0.77, height=20)

            delete = tk.Button(self, text="Delete", command = delete, width=10)
            delete.place(rely=0.02, relx=0.65, height=20)

        tv1.bind('<<TreeviewSelect>>', on_select) 

class optimize(GUI):
    def __init__(self, parent, controller):
        GUI.__init__(self, parent)

        frame1 = tk.LabelFrame(self, text="Product Details")
        frame1.place(rely=0.02, relx=0.02, height=480, width=520)

        frame2 = tk.LabelFrame(self, text="Shelf Details")
        frame2.place(rely=0.02, relx=0.68, height=480, width=250)

        shelf_info = tk.LabelFrame(frame2, text="Shelf Dimension")
        shelf_info.grid(row=0, column=0, sticky="news", padx=5, pady=5)

        shelf_dim_labelX = tk.Label(shelf_info, text="Length")
        shelf_dim_labelX.grid(row=0, column=0)
        shelf_dimL_spinbox = tk.Spinbox(shelf_info, from_=1, to=100)
        shelf_dimL_spinbox.grid(row=0, column=1)
        shelf_dim_labelY = tk.Label(shelf_info, text="Breadth")
        shelf_dim_labelY.grid(row=1, column=0)
        shelf_dimB_spinbox = tk.Spinbox(shelf_info, from_=1, to=100)
        shelf_dimB_spinbox.grid(row=1, column=1)
        shelf_dim_labelZ = tk.Label(shelf_info, text="Height")
        shelf_dim_labelZ.grid(row=2, column=0)
        shelf_dimH_spinbox = tk.Spinbox(shelf_info, from_=1, to=100)
        shelf_dimH_spinbox.grid(row=2, column=1)

        for widget in shelf_info.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        storage_info = tk.LabelFrame(frame2, text="Storage Info")
        storage_info.grid(row=2, column=0, sticky="news", padx=5, pady=5)

        no_of_shelf_refill = tk.Label(storage_info, text="Refill (Days)")
        no_of_shelf_refill.grid(row=0, column=0)
        no_of_shelf_refill_entry = tk.Entry(storage_info)
        no_of_shelf_refill_entry.grid(row=0, column=1)

        item = ttk.LabelFrame(frame2, text="Products to be placed on shelf")
        item.grid(row=3, columnspan=2, sticky="news", padx=5, pady=5)

        df = pd.read_excel('data.xlsx', sheet_name="category") 
        df1 = pd.read_excel('data.xlsx')

        def prinnt(selection):
            print(selection) 

        var = tk.StringVar(frame2)
        prod_category = ttk.OptionMenu(item, var, "Select Product", command=prinnt)
        prod_category.pack()

        menu = prod_category["menu"]

        vars = []           
        
        for n in df["Product Category"]:
            a=n
            n = tk.Menu(menu)
            menu.add_cascade(label = a, menu = n)
            for i in range(len(df1['product_name'])):
                if df1['category'][i] == a:
                    vars.append(df1['product_name'][i])
                    n.add_command(label=df1['product_name'][i])

        for widget in storage_info.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        blank1 = tk.Label(frame2)
        blank1.grid(row=1, columnspan=2)

        tv1 = ttk.Treeview(frame1)
        column_list = ["Name", "Volume", "Stock Ratio", "Ideal Stock (Month)", "Buffer Stock (Day)"]
        tv1['columns'] = column_list
        tv1["show"] = "headings"  # removes empty column
        for column in column_list:
            tv1.heading(column, text=column)
            tv1.column(column, anchor="center")
        tv1.column(0, width=1, anchor="w")
        tv1.column(1, width=5)
        tv1.column(2, width=5)
        tv1.column(3, width=5)
        tv1.column(4, width=5)
        tv1.place(relheight=1, relwidth=0.995)
        treescroll = tk.Scrollbar(frame1)
        treescroll.configure(command=tv1.yview)
        tv1.configure(yscrollcommand=treescroll.set)
        treescroll.pack(side="right", fill="y")

        def optimize():
            if no_of_shelf_refill_entry.get() != "" and shelf_dimL_spinbox.get() != "" and shelf_dimB_spinbox.get() != "" and shelf_dimH_spinbox.get() != "":
                value_check = [no_of_shelf_refill_entry.get(), shelf_dimL_spinbox.get(), shelf_dimB_spinbox.get(), shelf_dimH_spinbox.get()] 
                try:
                    [float(n) for n in value_check]
                    
                    df = pd.read_excel('data.xlsx')
                    products_names = df['product_name']
                    dimL = df['product_dimL']
                    dimB = df['product_dimB']
                    dimH = df['product_dimH']
                    sales_months = []
                    total_sales_list = []
                    for months in df.keys()[-3:]:
                        sales_months.append(months)
                        total_sales_list.append(df[months].sum())

                    shelf_volume = round(float(shelf_dimL_spinbox.get())*float(shelf_dimB_spinbox.get())*float(shelf_dimH_spinbox.get()),2)
                    shelf_refill = round(float(no_of_shelf_refill_entry.get()),2)
                    
                    blank1.config(text="Shelf Volume is " +str(shelf_volume)) 
                    frame1.config(text="Here are the Results")

                    tv1.delete(*tv1.get_children())
                    for n in range(len(products_names)):
                        name = products_names[n]
                        volume = round(dimL[n]*dimB[n]*dimH[n],2)
                        prod_sales = df[sales_months[0]][n]+df[sales_months[1]][n]+df[sales_months[2]][n]
                        total_sales = total_sales_list[0]+total_sales_list[1]+total_sales_list[2]

                        tv1.insert('','end', values=(name, volume, round(prod_sales/total_sales*100,2), math.ceil(round((prod_sales/(total_sales)*shelf_volume)/volume,2)), math.ceil((prod_sales/total_sales*shelf_volume)/round(volume,2)*shelf_refill)))
                   
                except(ValueError):
                    tv1.delete(*tv1.get_children())
                    frame1.config(text="Please Enter required format")  
            else:
                tv1.delete(*tv1.get_children())
                frame1.config(text="Please Enter required values")
            
        button = tk.Button(frame2, text="Optimize", command = optimize)
        button.grid(row=4, columnspan=2, sticky="news", padx=5, pady=5)

class pricing(GUI):
    def __init__(self, parent, controller):
        GUI.__init__(self, parent)

        frame1 = tk.LabelFrame(self, text="Profit Maximization")
        frame1.place(rely=0.02, relx=0.02, height=480, width=570)

        tv1 = ttk.Treeview(frame1)
        column_list = ["Name", "Selling Price", "Sales 3M", "Sales",  "Last 3M Profit", "Last Month Profit",""]
        tv1['columns'] = column_list
        tv1["show"] = "headings"  # removes empty column
        for column in column_list:
            tv1.heading(column, text=column)
            tv1.column(column, anchor="center")
        tv1.column(0, width=150, minwidth=150, anchor="w")
        tv1.column(1, width=80, minwidth=80)
        tv1.column(2, width=60, minwidth=60)
        tv1.column(3, width=60, minwidth=60)
        tv1.column(4, width=100, minwidth=100)
        tv1.column(5, width=100, minwidth=100)
        tv1.column(6, width=100, minwidth=50)
        tv1.place(relheight=1, relwidth=0.995)
        treescroll = tk.Scrollbar(frame1)
        treescroll.configure(command=tv1.yview)
        tv1.configure(yscrollcommand=treescroll.set)
        treescroll.pack(side="right", fill="y")

        def sort_data(tv, col, reverse):
            data = [(tv.set(child, col), child) for child in tv.get_children('')]
            if col == "Selling Price":
                data = [(int(val), child) for val, child in data]
            if col == "Sales 3M":
                data = [(int(val), child) for val, child in data]
            if col == "Sales":
                data = [(int(val), child) for val, child in data]
            if col == "Last 3M Profit":
                data = [(float(val), child) for val, child in data]
            if col == "Last Month Profit":
                data = [(float(val), child) for val, child in data]
            data.sort(reverse=reverse)
            for index, (val, child) in enumerate(data):
                tv.move(child, '', index)
            tv.heading(col, command=lambda: sort_data(tv, col, not reverse))
        
        df = pd.read_excel('data.xlsx')
        product_name = df['product_name']
        last3months = []
        for m in df.keys()[-3:]:
            last3months.append(m)

        for n in range(len(product_name)):                
            last3monthsSales= df[last3months[0]][n]+df[last3months[1]][n]+df[last3months[2]][n]
            profit3M = df[last3months[2]][n]*((float(df['product_SP'][n])-float(df['product_CP'][n])))+df[last3months[1]][n]*((float(df['product_SP'][n])-float(df['product_CP'][n])))+df[last3months[2]][n]*((float(df['product_SP'][n])-float(df['product_CP'][n])))
            tv1.insert('','end', values=(product_name[n], df['product_SP'][n], last3monthsSales, df[last3months[2]][n], profit3M, df[last3months[2]][n]*((float(df['product_SP'][n])-float(df['product_CP'][n])))))
        sort_data(tv1, "Name", True)
        sort_data(tv1, "Selling Price", True)
        sort_data(tv1, "Sales 3M", True)
        sort_data(tv1, "Sales", True)
        sort_data(tv1, "Last 3M Profit", True)
        sort_data(tv1, "Last Month Profit", True)

        def showResults():
            topShow = int(top_entry.get())
            tv1.delete(*tv1.get_children()[topShow:])            

        frame2 = tk.LabelFrame(self, text="Product Details")
        frame2.place(rely=0.02, relx=0.74, height=480, width=200)

        def option_selected(selection):
            tv1.delete(*tv1.get_children())
            df = pd.read_excel('data.xlsx')
            product_name = df['product_name']
            last3months = []
            for m in df.keys()[-3:]:
                last3months.append(m)
            if selection == "Select All":
                for n in range(len(product_name)):                
                    last3monthsSales= df[last3months[0]][n]+df[last3months[1]][n]+df[last3months[2]][n]
                    profit3M = df[last3months[2]][n]*((float(df['product_SP'][n])-float(df['product_CP'][n])))+df[last3months[1]][n]*((float(df['product_SP'][n])-float(df['product_CP'][n])))+df[last3months[2]][n]*((float(df['product_SP'][n])-float(df['product_CP'][n])))
                    tv1.insert('','end', values=(product_name[n], df['product_SP'][n], last3monthsSales, df[last3months[2]][n], profit3M, df[last3months[2]][n]*((float(df['product_SP'][n])-float(df['product_CP'][n])))))
                sort_data(tv1, "Name", True)
                sort_data(tv1, "Selling Price", True)
                sort_data(tv1, "Sales 3M", True)
                sort_data(tv1, "Sales", True)
                sort_data(tv1, "Last 3M Profit", True)
                sort_data(tv1, "Last Month Profit", True)
            
            else:
                df = pd.read_excel('data.xlsx')
                for i in range(len(product_name)):     
                    if selection == df['category'][i]:
                        last3monthsSales= df[last3months[0]][i]+df[last3months[1]][i]+df[last3months[2]][i]
                        profit3M = df[last3months[2]][i]*((float(df['product_SP'][i])-float(df['product_CP'][i])))+df[last3months[1]][i]*((float(df['product_SP'][i])-float(df['product_CP'][i])))+df[last3months[2]][i]*((float(df['product_SP'][i])-float(df['product_CP'][i])))
                        tv1.insert('','end', values=(product_name[i], df['product_SP'][i], last3monthsSales, df[last3months[2]][i], profit3M, df[last3months[2]][i]*((float(df['product_SP'][i])-float(df['product_CP'][i])))))
                    sort_data(tv1, "Name", True)
                    sort_data(tv1, "Selling Price", True)
                    sort_data(tv1, "Sales 3M", True)
                    sort_data(tv1, "Sales", True)
                    sort_data(tv1, "Last 3M Profit", True)
                    sort_data(tv1, "Last Month Profit", True)

        product_category = ["Select All"]
        value_inside = tk.StringVar(self)

        c = pd.read_excel('data.xlsx', sheet_name="category")
        for c in c['Product Category']:
            product_category.append(c)
        
        category = ttk.OptionMenu(frame2, value_inside, "Select Category", *product_category, command=option_selected)
        category.grid(row=0, columnspan=2, padx=10, pady=10, sticky="news")

        top = tk.Label(frame2, text="Products")
        top.grid(row=1, column=0, padx=10, pady=5, sticky="news")
        top_entry = tk.Entry(frame2)
        top_entry.grid(row=1, column=1, padx=10, pady=5, sticky="news")

        results = tk.Button(frame2, text="Show Result", command=showResults)
        results.grid(row=2, columnspan=2, padx=10, pady=5, sticky="news")

class Forecast(GUI):
    def __init__(self, parent, controller):
        GUI.__init__(self, parent)

        df = pd.read_excel('data.xlsx')

        frame1 = tk.LabelFrame(self.main_frame, text="Product")
        frame1.place(rely=0.02, relx=0.02, height=480, width=300)

        tv1 = ttk.Treeview(frame1)
        column_list = ["Sr no", "Name", "Profit"]
        tv1['columns'] = column_list
        tv1["show"] = "headings"  # removes empty column
        for column in column_list:
            tv1.heading(column, text=column)
            tv1.column(column, anchor="center")
        tv1.column(0, width=1)
        tv1.column(1, width=5, anchor="w")
        tv1.column(2, width=5)
        tv1.place(relheight=1, relwidth=0.995)
        treescroll = tk.Scrollbar(frame1)
        treescroll.configure(command=tv1.yview)
        tv1.configure(yscrollcommand=treescroll.set)
        treescroll.pack(side="right", fill="y")

        frame2 = tk.LabelFrame(self.main_frame, text="Sales")
        frame2.place(rely=0.02, relx=0.41, height=400, width=200)

        tv2 = ttk.Treeview(frame2)
        column_list = ["Month", "Sales"]
        tv2['columns'] = column_list
        tv2["show"] = "headings"  # removes empty column
        for column in column_list:
            tv2.heading(column, text=column)
            tv2.column(column, anchor="center")
        tv2.column(0, width=1)
        tv2.column(1, width=5)
        tv2.place(relheight=1, relwidth=0.995)
        treescroll = tk.Scrollbar(frame2)
        treescroll.configure(command=tv2.yview)
        tv2.configure(yscrollcommand=treescroll.set)
        treescroll.pack(side="right", fill="y")

        def Load_data():
            product_name = df['product_name']
            for n in range(len(product_name)):
                tv1.insert('','end', values=(n+1, product_name[n], float(df['product_SP'][n])-float(df['product_CP'][n])))

        frame3 = tk.LabelFrame(self.main_frame)
        frame3.place(rely=0.83, relx=0.41, height=75, width=200)

        frame3_label = tk.Label(frame3)
        frame3_label.grid(row=0, column=0)  

        frame4 = tk.LabelFrame(self.main_frame, text="Sales Earning Prediction")
        frame4.place(rely=0.02, relx=0.67, height=240, width=250)

        tv4 = ttk.Treeview(frame4)
        column_list = ["Month", "Earning Prediction"]
        tv4['columns'] = column_list
        tv4["show"] = "headings"  # removes empty column
        for column in column_list:
            tv4.heading(column, text=column)
            tv4.column(column, anchor="center")
        tv4.column(0, width=1)
        tv4.column(1, width=5)
        tv4.place(relheight=1, relwidth=0.995)
        treescroll = tk.Scrollbar(frame4)
        treescroll.configure(command=tv4.yview)
        tv4.configure(yscrollcommand=treescroll.set)
        treescroll.pack(side="right", fill="y")

        frame5 = tk.LabelFrame(self.main_frame, text="Sales Forecast")
        frame5.place(rely=0.5, relx=0.67, height=240, width=250)

        tv3 = ttk.Treeview(frame5)
        column_list = ["Month", "Sales Prediction"]
        tv3['columns'] = column_list
        tv3["show"] = "headings"  # removes empty column
        for column in column_list:
            tv3.heading(column, text=column)
            tv3.column(column, anchor="center")
        tv3.column(0, width=1)
        tv3.column(1, width=5)
        tv3.place(relheight=1, relwidth=0.995)
        treescroll = tk.Scrollbar(frame5)
        treescroll.configure(command=tv3.yview)
        tv3.configure(yscrollcommand=treescroll.set)
        treescroll.pack(side="right", fill="y")

        def on_select(event):

            selected_item = tv1.item(tv1.selection()[0], 'values')[1]
            selected_item_index = int(tv1.item(tv1.selection()[0], 'values')[0]) - 1
            product_sales = {}
            for n in range(len(df.keys()) - 7):
                key = str(df.keys()[7+n])[14:]
                value = df['product_sales_'+ key][selected_item_index]
                product_sales.update({key: value})
            
            product_sales_months = list(product_sales.keys())
            product_sales_sales = list(product_sales.values())

            frame2.config(text="Sales of" + " " +selected_item)
            tv2.delete(*tv2.get_children())

            for i in range(len(product_sales)):
                tv2.insert('','end', values=(product_sales_months[i], product_sales_sales[i]))

            def salesMonth():
                plt.clf()
                plt.plot(product_sales_months[-12:], product_sales_sales[-12:])
                plt.xlabel('Month -->', fontsize=10)
                plt.xticks(product_sales_months[-12:], fontsize=7)
                plt.ylabel('Product Sales -->', fontsize=10)
                plt.yticks(product_sales_sales[-12:], fontsize=7)
                plt.title("Sales Per Month"+"("+selected_item+")")
                plt.show()

            def forecast():
                d = product_sales_sales[-12:]

                model_sarima=sm.tsa.statespace.SARIMAX(d, order=(1,1,1), seasonal_order=(1,0,1,4))
                results=model_sarima.fit()

                f_months = []
                current_month= str(product_sales_months[-1])[:-4]
                current_year=str(product_sales_months[-1])[-4:]
                
                for i in range(6):
                    f_months.append(months[months.index(current_month)+1+i]+current_year)
                    if months.index(current_month)+2+i>11:
                        current_year = current_year +1

                df_forecast = results.predict(start = 12, end = 17).tolist()
                df_forecast = [round(n, 2) for n in df_forecast]

                plt.clf()
                plt.plot(product_sales_months[-12:], d)
                plt.plot(f_months, df_forecast)
                plt.xlabel('Month -->', fontsize=10)
                plt.xticks(product_sales_months[-12:] + f_months, fontsize=7)
                plt.ylabel('Product Sales -->', fontsize=10)
                plt.yticks(product_sales_sales[-12:] + df_forecast, fontsize=7)
                plt.title("Sales Per Month"+"("+selected_item+")")
                plt.legend(["Sales", "Forecast"])
                plt.show()

                frame5.config(text="Sales Forecast of" + " " +selected_item)
                tv3.delete(*tv3.get_children())
                for i in range(len(f_months)):
                    tv3.insert('','end', values=(f_months[i], df_forecast[i]))

                df = pd.read_excel('data.xlsx')
                frame4.config(text="Sales Earning Prediction" + " " +selected_item)
                tv4.delete(*tv4.get_children())
                for i in range(len(f_months)):
                    tv4.insert('','end', values=(f_months[i], round(df_forecast[i]*float(df['product_SP'][i]-df['product_CP'][i]),2)))

            graph1 = tk.Button(frame3_label, text="Sales/Month", command = salesMonth, width=24)
            graph1.grid(row=0, column=0, sticky="news", padx=5, pady=5)

            graph2 = tk.Button(frame3_label, text="forecast", command = forecast, width=24)
            graph2.grid(row=1, column=0, sticky="news", padx=5, pady=5)

        tv1.bind('<<TreeviewSelect>>', on_select)

        Load_data()

class About(GUI):
    def __init__(self, parent, controller):
        GUI.__init__(self, parent)

        label1 = tk.Label(self.main_frame, font=("Verdana", 15), text="Developed by Shobhit & Shruti Juyal")
        label1.pack(side="top")

class OpenNewWindow(tk.Tk):
    def __init__(self, *args, **kwargs):

        tk.Tk.__init__(self, *args, **kwargs)

        main_frame = tk.Frame(self)
        main_frame.pack_propagate(0)
        main_frame.pack(fill="both", expand="true")
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)
        self.title("Here is the Title of the Window")
        self.geometry("500x500")
        self.resizable(0, 0)

        frame1 = ttk.LabelFrame(main_frame, text="This is a ttk LabelFrame")
        frame1.pack(expand=True, fill="both")

        label1 = tk.Label(frame1, font=("Verdana", 20), text="OpenNewWindow Page")
        label1.pack(side="top")

top = LoginPage()
top.title("Login")
root = MyApp()
root.withdraw()
root.title("Shelf Optimization")

root.mainloop()